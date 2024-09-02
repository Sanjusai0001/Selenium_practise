import time

import openpyxl

# file = "C:/Users/sanju/PycharmProjects/SELENIUM/test_data.xlsx"
file = "C:/Users/sanju/PycharmProjects/SELENIUM/test_data1.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook.active # (OR)  sheet = workbook[Data] -- get active sheet from excel


# for r in range(1,5):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Hello"  # writing
#
# workbook.save(file)


sheet.cell(1,1).value= 1
sheet.cell(1,2).value="Ram"
sheet.cell(1,3).value="Developer"
sheet.cell(1,4).value="NYC"

sheet.cell(2,1).value= 2
sheet.cell(2,2).value="Sam"
sheet.cell(2,3).value="System Engineer"

sheet.cell(3,1).value= 3
sheet.cell(3,2).value="Prem"
sheet.cell(3,3).value="Advertiser"
sheet.cell(3,4).value="UK"

sheet.cell(4,1).value= 4
sheet.cell(4,2).value="Zara"
sheet.cell(4,3).value="Tester"

sheet.cell(5,1).value= 5
sheet.cell(5,2).value="Gaara"
sheet.cell(5,3).value="AI/ML Engineer"

sheet.cell(6,1).value= 6
sheet.cell(6,3).value="Bran"
sheet.cell(8,2).value="Data Engineer"


curr = time.ctime(1721029393.0197895)
print("Current time:", curr)
workbook.save(file)