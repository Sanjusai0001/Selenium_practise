import openpyxl


# File --> Workbook --> Sheets --> Rows --> Cells


# path
# file = "C:\Users\sanju\PycharmProjects\SELENIUM\sample_data.xlsx"
# file = "//SELENIUM/sample_data.xlsx"
# file =  "C:\Users\sanju\PycharmProjects\SELENIUM\sample_data.xlsx"
file = "C:/Users/sanju/PycharmProjects/SELENIUM/sample_data.xlsx"

workbook = openpyxl.load_workbook(file)
sheet=workbook['Sheet1']

rows = sheet.max_row  # count no.of rows in excel sheet
cols = sheet.max_column # count no.of column in excel sheet


# Reading all the rows & columns from excel sheet

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end='     -     ')
    print()